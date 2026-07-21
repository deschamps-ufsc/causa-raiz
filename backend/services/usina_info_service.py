"""
Serviço de armazenamento de Infos Usina (tabela Série × Qtde Módulos × Wp).
Salva/lê DATA_DIR/{usina}/usina_info.json
"""
import io
import json
import os

import pandas as pd

from utils.config import DATA_DIR
from utils.logger import logger


def _info_file(usina: str) -> str:
    return os.path.join(DATA_DIR, usina, "usina_info.json")


# ── Carregar / Salvar ─────────────────────────────────────────────────────────

def load_usina_info(usina: str) -> dict:
    """Retorna dict {serie_temporal: {qtde_modulos, wp, kwp}}."""
    path = _info_file(usina)
    if not os.path.exists(path):
        return {}
    with open(path, "r", encoding="utf-8") as f:
        return json.load(f)


def save_usina_info(usina: str, records: dict) -> None:
    """Persiste o dicinario em disco."""
    path = _info_file(usina)
    os.makedirs(os.path.dirname(path), exist_ok=True)
    with open(path, "w", encoding="utf-8") as f:
        json.dump(records, f, indent=2, ensure_ascii=False)
    logger.info(f"[USINA_INFO] Salvo: {len(records)} séries -> {path}")


# ── Importar Excel ────────────────────────────────────────────────────────────

def import_info_from_excel(content: bytes, usina: str) -> dict:
    """
    Lê Excel com colunas:
        skid | inversor | stringbox | string | Qtde Módulos | Wp
    """
    df = pd.read_excel(io.BytesIO(content), engine="openpyxl")

    col_map = {}
    for c in df.columns:
        norm = c.strip().lower().replace(" ", "_").replace("é", "e").replace("ó", "o")
        col_map[c] = norm
    df.rename(columns=col_map, inplace=True)

    def find_col(keywords: list[str]) -> str | None:
        for kw in keywords:
            matches = [c for c in df.columns if kw in c]
            if matches:
                return matches[0]
        return None

    col_skid     = find_col(["skid"])
    col_inversor = find_col(["inversor", "inv"])
    col_sbox     = find_col(["stringbox", "sb"])
    
    # Custom exact match or exclusion to prevent "stringbox" matching "string"
    col_string = None
    for c in df.columns:
        if "string" in c and "box" not in c:
            col_string = c
            break
    
    col_qtde  = find_col(["qtde", "quantidade", "modulo"])
    col_wp    = find_col(["wp", "potencia"])

    if not col_skid or not col_wp:
        raise ValueError(
            "O Excel deve ter pelo menos as colunas: skid, Wp. "
            f"Colunas encontradas: {list(df.columns)}"
        )

    records = {}
    invalidas = 0
    for _, row in df.iterrows():
        def clean_val(v):
            if pd.isna(v) or str(v).lower() in ("nan", ""): return ""
            val_str = str(v).strip()
            if val_str.endswith(".0"): return val_str[:-2]
            return val_str

        skid = clean_val(row.get(col_skid, ""))
        inv = clean_val(row.get(col_inversor, "")) if col_inversor else ""
        sb = clean_val(row.get(col_sbox, "")) if col_sbox else ""
        st = clean_val(row.get(col_string, "")) if col_string else ""

        if not skid:
            invalidas += 1
            continue

        # Composite Key construction: skid|inversor|stringbox|string (falling back dynamically)
        parts = [p for p in [skid, inv, sb, st] if p]
        key = "|".join(parts)

        qtde = 1
        if col_qtde:
            try:
                raw_qtde = row.get(col_qtde)
                if pd.notna(raw_qtde) and str(raw_qtde).lower() != "nan":
                    qtde = int(float(raw_qtde))
            except (ValueError, TypeError):
                pass

        wp = None
        if col_wp:
            try:
                wp_raw = row.get(col_wp)
                if pd.notna(wp_raw) and str(wp_raw).lower() != "nan":
                    wp = float(wp_raw)
            except (ValueError, TypeError):
                pass
                
        if wp is None:
            invalidas += 1
            continue
            
        kwp = round((qtde * wp) / 1000.0, 4)

        records[key] = {
            "skid": skid,
            "inversor": inv,
            "stringbox": sb,
            "string": st,
            "qtde_modulos": qtde,
            "wp": wp,
            "kwp": kwp,
        }

    save_usina_info(usina, records)

    return {
        "total_series": len(records),
        "linhas_invalidas": invalidas,
    }


# ── Gerar template Excel ──────────────────────────────────────────────────────

def generate_info_template() -> bytes:
    """Gera um Excel modelo para o usuário preencher."""
    data = {
        "skid": ["CLS01.1", "CLS01.1"],
        "inversor": ["INV01", "INV01"],
        "stringbox": ["SB13", "SB13"],
        "string": [1, 2],
        "Qtde Módulos": [31, 31],
        "Wp": [615, 615],
    }
    df = pd.DataFrame(data)

    output = io.BytesIO()
    with pd.ExcelWriter(output, engine="openpyxl") as writer:
        df.to_excel(writer, index=False, sheet_name="Infos Usina")
        ws = writer.sheets["Infos Usina"]
        from openpyxl.styles import Font, PatternFill
        for cell in ws[1]:
            cell.font = Font(bold=True, color="FFFFFF")
            cell.fill = PatternFill("solid", fgColor="1F2937")
        for col in ws.columns:
            max_len = max(len(str(cell.value or "")) for cell in col)
            ws.column_dimensions[col[0].column_letter].width = max(max_len + 4, 30)

    return output.getvalue()
